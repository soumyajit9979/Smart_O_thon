import serial
import datetime
import json
import openpyxl
import os

# Establish serial connection with Arduino
ser = serial.Serial('COM3', 9600)  # Replace 'COM4' with the appropriate serial port

attendance_data = []

while True:
    # Read data from Arduino
    data = ser.readline().decode().strip()
    if data:
        fingerprint_id = data.split()[0]
        print(fingerprint_id)

        if fingerprint_id == "12":
            fingerprint_id = "OM"
        elif fingerprint_id == "56":
            fingerprint_id = "soumyajit"
        elif fingerprint_id == "4":
            fingerprint_id = "Teacher locked the attendance"
            # exit(0)
        # elif fingerprint_id == "55":
        #     fingerprint_id = "Unlocked by Bluetooth"
        else:
            fingerprint_id = "invalid"

        # Get the current timestamp
        timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Create a dictionary to store the fingerprint data
        fingerprint = {"id": fingerprint_id, "timestamp": timestamp}

        # Append the fingerprint data to the attendance list
        attendance_data.append(fingerprint)

        # Save the attendance data to a JSON file
        with open("fingerprint_data.json", "w") as json_file:
            json.dump(attendance_data, json_file)

        # Save the attendance data to an Excel file
        if os.path.isfile("fingerprint_data.xlsx"):
            workbook = openpyxl.load_workbook("fingerprint_data.xlsx")
        else:
            workbook = openpyxl.Workbook()
            workbook.active['A1'] = "ID"
            workbook.active['B1'] = "Name"
            workbook.active['C1'] = "Timestamp"

        sheet = workbook.active
        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=fingerprint_id)
        sheet.cell(row=row, column=2, value=fingerprint_id)
        sheet.cell(row=row, column=3, value=timestamp)
        # sheet.cell(row=row, column=4, value="PRESENT")

        workbook.save("fingerprint_data.xlsx")

        if(fingerprint_id == "Teacher locked the attendance"):
            exit(0)
